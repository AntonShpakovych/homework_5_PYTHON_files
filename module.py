import json
import csv
from os import path, sep
from openpyxl import Workbook
from openpyxl import load_workbook

"""Module with func"""


# -------JSON------------------

def write_json(path_file, data):
    with open(path_file, 'w') as outfile:
        json.dump(data, outfile, indent=4, sort_keys=True, ensure_ascii=True)
    outfile.close()


def read_json(path_file):
    file = open(path_file, 'r')
    data = json.loads(file.read())
    for item in data['exchangeRate']:
        print(item)
    file.close()


# ----------------CSV--------------------------
def write_csv(path_file, data):
    header = list(data['exchangeRate'][0].keys())
    with open(path_file, 'w', encoding='UTF8', newline='') as file:
        writer = csv.DictWriter(file, fieldnames=header)
        writer.writeheader()
        writer.writerows(data['exchangeRate'])

    file.close()


def read_csv(path_file):
    with open(path_file, 'r', encoding='UTF8', newline='') as file:
        filereader = csv.DictReader(file)
        for row in filereader:
            print(row)
    file.close()


# -------------XLSX------------------------


def write_xlsx(path_name, data):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'PB_data'
    for index, item in enumerate(list(data['exchangeRate'][0].keys())):
        sheet.cell(row=1, column=index+1).value = item

    row_num = 1
    for row in data['exchangeRate']:
        row_num += 1
        for index, key in enumerate(list(row.keys())):
            sheet.cell(row=row_num, column=index+1).value = row[key]

    workbook.save(filename="./file/file_3.xlsx")


def read_xlsx(path_name):
    workbook = load_workbook(filename=path_name)
    workbook.sheetnames
    sheet = workbook.active
    data = {}
    data['exchangeRate'] = []
    keys = []
    for value in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=6, values_only=True):
        for id, key in enumerate(value):
            keys.append(key)

    for values in sheet.iter_rows(min_row=2, max_row=4, min_col=1, max_col=6, values_only=True):
        row = {}
        for id, value in enumerate(values):
            row.update({keys[id]: value})
        data['exchangeRate'].append(row)

    print(data)
